import tkinter
from PIL import ImageTk, Image
from tkinter import Tk, Label, Button,Entry, Frame, END , messagebox ,ttk
from tkinter import *
from tkinter.ttk import *
import cv2
import os
import imutils
import datetime
import time
import numpy as np
import pandas as pd
import openpyxl
import serial
import datetime
from openpyxl import load_workbook
from openpyxl.chart import Reference,  LineChart
from git import Repo
from datetime import date
import PIL.Image
import PIL.ImageTk
import paho.mqtt.client as mqtt
import mariadb

dato8 = 0
dato9 = 0

def showMessage(message, type='info', timeout=2500):
    from tkinter import Tk
    from tkinter.messagebox import Message 
    from _tkinter import TclError

    import tkinter as tk
    from tkinter import messagebox as msgb

    root = tk.Tk()
    root.withdraw()
    try:
        root.after(timeout, root.destroy)
        if type == 'info':
            msgb.showinfo('Info', message, master=root)
        elif type == 'warning':
            msgb.showwarning('Warning', message, master=root)
        elif type == 'error':
            msgb.showerror('Error', message, master=root)
    except:
        pass
  
def inicio():
    tvX1=40; tvX2=190
    etX1=130; etY1=30
    btX1=200; btY1=50

    Temperaturaz=0.0
    Peso=0.0
    Estatura=0.0
    Imc=0.0
    O2Sat=0.0
 
    ventIoMT = Toplevel(ventana)
    ventIoMT.title("IoMT")
    ventIoMT.geometry('%dx%d+%d+%d' % (360, 450, 51, 60))
    ventIoMT.resizable(0, 0)

    # VIZUALIZACION DE RESULTADO
    tv1Temperatura= tkinter.Label(ventIoMT, text= str("Temperatura"), font=("Times New Roman",20))
    tv1Temperatura.place(x=tvX1, y=100)
    tv1Peso= tkinter.Label(ventIoMT, text= str("Peso"), font=("Times New Roman",20))
    tv1Peso.place(x=tvX1, y=160)
    tv1Estatura= tkinter.Label(ventIoMT, text= str("Estatura"), font=("Times New Roman",20))
    tv1Estatura.place(x=tvX1, y=220)
    tv1Imc= tkinter.Label(ventIoMT, text= str("IMC"), font=("Times New Roman",20))
    tv1Imc.place(x=tvX1, y=280)
    tv1O2Sat= tkinter.Label(ventIoMT, text= str("O2Sat"), font=("Times New Roman",20))
    tv1O2Sat.place(x=tvX1, y=340)

    # VIZUALIZACION DE RESULTADO
    tv2Temperatura= tkinter.Label(ventIoMT, text= str(Temperaturaz), font=("Times New Roman",20),bg='#fff')
    tv2Temperatura.place(x=tvX2, y=100, width=etX1, height=etY1)
    tv2Peso= tkinter.Label(ventIoMT, text= str(Peso), font=("Times New Roman",20),bg='#fff')
    tv2Peso.place(x=tvX2, y=160, width=etX1, height=etY1)
    tv2Estatura= tkinter.Label(ventIoMT, text= str(Estatura), font=("Times New Roman",20),bg='#fff')
    tv2Estatura.place(x=tvX2, y=220, width=etX1, height=etY1)
    tv2Imc= tkinter.Label(ventIoMT, text= str(Imc), font=("Times New Roman",20),bg='#fff')
    tv2Imc.place(x=tvX2, y=280, width=etX1, height=etY1)
    tv2O2Sat= tkinter.Label(ventIoMT, text= str(O2Sat), font=("Times New Roman",20),bg='#fff')
    tv2O2Sat.place(x=tvX2, y=340, width=etX1, height=etY1)
    
    showMessage('La informacion se enviará durante 15 segundos', type='info', timeout=2000)
    contador =0
    serialArduino = serial.Serial('/dev/ttyACM0',9600)
    contadorprogres=0
    
    while True:
        comando='O'
        comandoBytes = comando.encode()
        serialArduino .write(comandoBytes)       
        
        progress['value'] = contadorprogres
        ventana.update_idletasks()
        txt2['text']=round(progress['value']),'%'
        cad=serialArduino.readline().decode('ascii')
        datoss=cad.splitlines()
        d0=str(datoss[0])
        d1=d0.replace("b","")
        d2=d1.replace("'","")
        d3=d2.split(",")
        dato1=str(d3[0])
        dato2=str(d3[1])
        dato3=str(d3[2])
        dato4=str(d3[3])
        dato5=str(d3[4])
                    
        dato9=float(dato3)
        if(dato9<1):
            dato9=1
                        
        dato6=float(dato2)/(float(dato9)*float(dato9))
        dato7=round(dato6,2)
        dato8=str(dato7)
        
        tv2Temperatura['text']=dato1
        tv2Peso['text']=dato2
        tv2Estatura['text']=dato3
        tv2Imc['text']=dato8
        tv2O2Sat['text']=dato5
        
        print(dato1,dato2,dato3,dato8,dato5)
        print("--------------------")
        print(contador)
        print("--------------------")
        client = mqtt.Client()
        client.connect("test.mosquitto.org",1883,60)
        client.publish("Topic1/Topic2/Topic3", str(dato1)+';'+str(dato2)+';'+str(dato3)+';'+str(dato8)+';'+str(dato5));
        client.disconnect();
        contador=contador+1
        contadorprogres=((contador*100)/14)
        
        try :
            if contador>= 15:
                showMessage('IOMT realizado Exitosamente', type='info', timeout=2000)
                comando='F'
                comandoBytes = comando.encode()
                serialArduino .write(comandoBytes)
                serialArduino.close()
                progress['value'] = 0
                txt2['text']=round(progress['value']),'%'
                ventIoMT.destroy()
                break
            
        except IndexError:
            messagebox.showinfo('IoMT','Vuelva a Intentarlo')
 
def ventanaregis():
    ventana.withdraw()
    ventana2 = tkinter.Toplevel()
    ventana2.attributes('-fullscreen', True)
    ventana2.title('REGISTRO')
    labelregistro = Label(ventana2, image=photo1).place(x=0,y=0,relwidth=1.0,relheight=1.0)

    entradanombre= tkinter.Entry(ventana2, text= 'Nombre',font=('ALGERIAN 20 bold'))
    entradanombre.place(x=460 , y=126 , width=652, height=70)
    entradaapellido= tkinter.Entry(ventana2, text= 'Apellido',font=('ALGERIAN 20 bold'))
    entradaapellido.place(x=460 , y=210 , width=652, height=70)
    entradacedula= tkinter.Entry(ventana2, text= 'Cédula',font=('ALGERIAN 20 bold'))
    entradacedula.place(x=460 , y=294 , width=652, height=70)
    entradacurso = ttk.Combobox(ventana2, state = 'readonly',font=('ALGERIAN 20 bold'))
    entradacurso['values']=['Seleccione','Inicial 1','Inicial 2','Primero','Segundo','Tercero','Cuarto','Quinto','Sexto','Séptimo','Octavo','Noveno','Décimo','Primero de Bachillerato','Segundo de Bachillerato','Tercero de Bachillerato','Básica Acelerado','Docente','Otros']
    entradacurso.current(0)
    entradacurso.place(x=460 , y=378 , width=652, height=70)
    
    entradaaño = ttk.Combobox(ventana2, state = 'readonly',font=('ALGERIAN 20 bold'))
    entradaaño['values']=['AÑO','2022','2021','2020','2019','2018','2017','2016','2015','2014','2013','2012','2011','2010','2009','2008','2007','2006','2005','2004','2003','2002','2001','2000','1999','1998','1997','1996','1995','1994','1993','1992','1991','1990','1989','1988','1987','1986','1985','1984','1983','1982','1981','1980','1979','1978','1977','1976','1975','1975','1974','1973','1972','1971','1970','1969','1968','1967','1966','1965','1964','1963','1962','1961','1960','1959','1958','1957','1956','1955','1954','1953','1952','1951','1950','1949','1948','1947','1946','1945','1944','1943','1942','1941','1940']
    entradaaño.current(0)
    entradaaño.place(x=508 , y=460 , width=170, height=70)
    
    entradames = ttk.Combobox(ventana2, state = 'readonly',font=('ALGERIAN 20 bold'))
    entradames['values']=['MES','1','2','3','4','5','6','7','8','9','10','11','12']
    entradames.current(0)
    entradames.place(x=698 , y=460 , width=170, height=70)
    
    entradadia = ttk.Combobox(ventana2, state = 'readonly',font=('ALGERIAN 20 bold'))
    entradadia['values']=['DIA','1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
    entradadia.current(0)
    entradadia.place(x=889 , y=460 , width=170, height=70)
    
     def entrenar():      
        dataPath = '/home/pi/IoMTv2/Data' #Cambia a la ruta donde hayas almacenado Data
        peopleList = os.listdir(dataPath)
        print('Lista de personas: ', peopleList)

        labels = []
        facesData = []
        label = 0

        for nameDir in peopleList:
            personPath = dataPath + '/' + nameDir
            print('Leyendo las imágenes')
            for fileName in os.listdir(personPath):
                print('Rostros: ', nameDir + '/' + fileName)
                labels.append(label)
                facesData.append(cv2.imread(personPath+'/'+fileName,0))
            label = label + 1
        face_recognizer = cv2.face.EigenFaceRecognizer_create()
        print("Entrenando...")
        face_recognizer.train(facesData, np.array(labels))
        face_recognizer.write('modeloEigenFacepruebainterface.xml')
        print("Modelo almacenado...")

    def mensajeregistro():
        messagebox.showinfo('Registro','Registro Realizado Correctamente')
    def mensajeerror():
        messagebox.showinfo('Error','El usuario que desea Registrar ya existe en la base de datos')  
    def atras():
        ventana2.withdraw()
        ventana.deiconify()            
    def calcular_edad_agnios(fecha_nacimiento):
        fecha_actual = date.today()
        resultadoaños = fecha_actual.year - fecha_nacimiento.year
        resultadoaños -= ((fecha_actual.month, fecha_actual.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
        return resultadoaños

    def capturarrostro():
        nombre = entradanombre.get()
        apellido = entradaapellido.get()
        cedula = entradacedula.get()
        curso = entradacurso.get()
        año = int(entradaaño.get())
        mes = int(entradames.get())
        dia = int(entradadia.get())        
        fecha_nacimiento_persona = date(año,mes,dia)      
        edad = str(calcular_edad_agnios(fecha_nacimiento_persona))
        nacimiento = (str(str(año)+"/"+str(mes)+"/"+str(dia)))
        
        if(entradacurso.get()== 'Seleccione'):
            curso = 'Otros'

        personName = nombre + ' ' + apellido
        dataPath = '/home/pi/IoMTv2/Data' #Cambia a la ruta donde hayas almacenado Data
        personPath = dataPath + '/' + personName
        
        if (os.path.isfile('/home/pi/IoMTv2/Todos los Datos/'+ personName +'.csv')):
            mensajeerror()     
        else:
            os.path.exists(personPath)
            print('Carpeta creada: ',personPath)
            os.makedirs(personPath)        
            cap = cv2.VideoCapture(0,cv2.CAP_V4L)
            faceClassif = cv2.CascadeClassifier(cv2.data.haarcascades+'haarcascade_frontalface_default.xml')
            count = 0
            countcerrar=0

            while True:
                ret, frame = cap.read()
                if ret == False: break
                frame =  imutils.resize(frame, width=640)
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                auxFrame = frame.copy()
                faces = faceClassif.detectMultiScale(gray,1.3,5)
                for (x,y,w,h) in faces:
                    cv2.rectangle(frame, (x,y),(x+w,y+h),(0,255,0),2)
                    rostro = auxFrame[y:y+h,x:x+w]
                    rostro = cv2.resize(rostro,(150,150),interpolation=cv2.INTER_CUBIC)
                    cv2.imwrite(personPath + '/rostro_{}.jpg'.format(count),rostro)
                    count = count + 1
                    countcerrar= countcerrar+1
                cv2.imshow('frame',frame)
                
                k =  cv2.waitKey(1)
                if k == 27 or count >= 20:
                
                    break
                if countcerrar>=50:
                    break
            cap.release()
            cv2.destroyAllWindows()           
            entrenar()           
            archivo = open ('/home/pi/IoMTv2/Todos los Datos/'+ personName +'.csv','a')
            archivo.write('Nombre '+ ','+ personName + '\n' + 'Cedula' + ',' + cedula +'\n'+ 'Curso' + ',' + curso + '\n'+'Fecha de Nacimiento' + ',' + nacimiento +'\n'+'Edad' + ',' + edad +'\n')
            archivo.close()          
            archivo1 = open ('/home/pi/IoMTv2/daniels36999.github.io/Datos Almacenados/'+curso+'/'+ personName +'.html','a')
            archivo1.write(
            """
            <!doctype html>
            <html lang="es">
                <head>
                    <meta charset="UTF-8">
                    <meta name="viewport"content="width=device-width, user-scalable=no, initial-scale=1.0, maximum-scale=1.0, minimum-scale=1.0">
                    <meta http-equiv="X-UA-Compatible" content="ie=edge">	
                    <title>HISTORIAL IoMT</title>	
                    <style type="text/css">#oli  {display: none;}</style>
                    <script   src="https://code.jquery.com/jquery-2.2.4.min.js"integrity="sha256-BbhdlvQf/xTY9gja0Dq3HiwQF8LaCRTXxZKRutelT44="   crossorigin="anonymous"></script>
                   <!-- LIBRERIAS EXCEL-->
                    <script src="https://unpkg.com/xlsx@0.16.9/dist/xlsx.full.min.js"></script>
                    <script src="https://unpkg.com/file-saverjs@latest/FileSaver.min.js"></script>
                    <script src="https://unpkg.com/tableexport@latest/dist/js/tableexport.min.js"></script>
                    <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/2.8.0/Chart.js"></script>
                    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.0.0-beta2/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-BmbxuPwQa2lc/FVzBcNJ7UAyJxM6wuqIj61tLrc4wSX0szH/Ev+nYRRuWlolflfl" crossorigin="anonymous">

                </head>

                <body>
                    <div class="container">
                        <div class="card" style="background:#E4FFB5  ;">
                            <div class="card-header" style="width: 27rem;">
                                REPORTE IOMT
                            </div>
                            <div class="card-body">
                                <button id="btnExportar" class="btn btn-success">
                                    <i class="fas fa-file-excel"></i> DESCARGAR HISTORIAL
                                </button>
                                <h4>
                                <table id="tabla" class="table table-border table-hover">
                                    <thead>
                                        <tr><th>Nombre:</th><th>"""+nombre+"""</th></tr>				
                                        <tr><th>Apellido:</th><th>"""+apellido+"""</th></tr>
                                        <tr><th>Cedula:</th><th>"""+cedula+"""</th></tr>
                                        <tr><th>Curso:</th><th>"""+curso+"""</th></tr>
                                        <tr><th>Fecha de Nacimiento:</th><th>"""+nacimiento+"""</th></tr>
                                        <tr><th>Edad:</th><th>"""+edad+"""</th></tr>
                                        <tr><th></th><th></th></tr>				
                                    
                                        <tr><th>Fecha</th>
                                        <th>Hora</th>
                                        <th>Temperatura</th>
                                        <th>Peso</th>
                                        <th>Altura</th>
                                        <th>IMC</th>
                                        <th>O2Sat</th></tr>
                                    </thead>
                                    <tbody>
                                    <tr><td class="GF">01/01/22</td><td class="GH">00:00</td><td class="GT">00.00</td><td class="GP">00.00</td><td class="GE">00.00</td><td class="GI">00.00</td><td class="GO">00.00</td></tr>
                                    <!-- AGREGAR TABLA-->  
                                        
                                    </tbody>
                                </table>
                            <!-- BOTON VISUALIZAR-->
                                <button id="btnGraficar" class="btn btn-success">
                                    <i class="fas fa-file-excel"></i> VISUALIZAR GRAFICAS
                                </button>
                                              
                                <div id="oli" >
                                    <!-- GRAFICA TEMPERATURA CORPORAL-->
                                    <canvas id="GraTemperatura" width="1000" height="500" style="margin:0 auto;"></canvas>
                                    <button id="DescTemp" class="btn btn-success"><i class="fas fa-file-excel"></i> DESCARGAR</button>
                                    
                                    <!-- GRAFICA PESO CORPORAL-->
                                    <canvas id="GraPeso" width="1000" height="400" style="margin:0 auto;"></canvas>
                                    <button id="DescPes" class="btn btn-success"><i class="fas fa-file-excel"></i> DESCARGAR</button>				
                                
                                    <!-- GRAFICA ESTATURA CORPORAL-->		
                                    <canvas id="GraEstatura" width="1000" height="400" style="margin:0 auto;"></canvas>				
                                    <button id="DescEst" class="btn btn-success"><i class="fas fa-file-excel"></i> DESCARGAR</button>
                                    
                                    <!-- GRAFICA IMC CORPORAL-->
                                    <canvas id="GraImc" width="1000" height="400" style="margin:0 auto;"></canvas>
                                    <button id="DescImc" class="btn btn-success"><i class="fas fa-file-excel"></i> DESCARGAR</button>
                                    
                                    <!-- GRAFICA O2SAT CORPORAL-->
                                    <canvas id="GraO2Sat" width="1000" height="400" style="margin:0 auto;"></canvas>
                                    <button id="DescO2Sat" class="btn btn-success" ><i class="fas fa-file-excel"></i> DESCARGAR</button>
                                    
                                    <!-- BOTON EXPORAR PDF-->
                                    <button type="button" class="btn btn-primary btn-sm" id="DescPDF"onclick="DescPDF">Descargar Historial en PDF</button>					
                            
                            <h4>
                            </div>
                        </div>
                    </div>
                    
                </body>
                
                <script src="https://ajax.googleapis.com/ajax/libs/jquery/2.1.1/jquery.min.js"></script>
                <script src='https://cdnjs.cloudflare.com/ajax/libs/jspdf/1.3.3/jspdf.debug.js'></script>
                        
                <!-- SCRIPT -------------------------------------------------------------------------- -->		
                <!-- SCRIPT PARA EXPORTAR -->
                <script>
                    const $btnExportar = document.querySelector("#btnExportar"),
                        $tabla = document.querySelector("#tabla");
                    $btnExportar.addEventListener("click", function() {
                        let tableExport = new TableExport($tabla, {
                            exportButtons: false, // No queremos botones
                            filename: "Reporte IoMT", //Nombre del archivo de Excel
                            sheetname: "IoMT", //Título de la hoja
                        });
                        let datos = tableExport.getExportData();
                        let preferenciasDocumento = datos.tabla.xlsx;
                        tableExport.export2file(preferenciasDocumento.data, preferenciasDocumento.mimeType, preferenciasDocumento.filename, preferenciasDocumento.fileExtension, preferenciasDocumento.merges, preferenciasDocumento.RTL, preferenciasDocumento.sheetname);
                    });
                </script>		
                        


                <!-- GRAFICA TEMPERATURA -->
                <!-- --------------------------------------------------------- -->
                <script>
                    var Temp = "";       var Fec = "";
                    var Temperatura =[]; var Fecha =[];
                    $("#btnGraficar").click(function() {
                        console.log("DESCARGANDOooo")
                    $(".GF").parent("tr").find(".GF").each(function() {
                     Fec +=$(this).html()+",";
                    });
                    $(".GT").parent("tr").find(".GT").each(function() {
                     Temp +=$(this).html()+",";
                    });
                    
                    Temp=Temp.split(',')
                    TempLen = Temp.length-1;
                    Fec=Fec.split(',')
                    for ( var i = 0; i < TempLen; i++ ) {
                        Temperatura[i]=Temp[i];
                        Fecha[i]=Fec[i];
                        }
                    var ctx= document.getElementById("GraTemperatura").getContext("2d");
                    let gradient=ctx.createLinearGradient(0,0,0,400);
                    gradient.addColorStop(0,"rgba(144,0,0,0.7)");
                    gradient.addColorStop(1,"rgba(200,0,0,0.3)");
                    var myChart= new Chart(ctx,{
                        type:"line", //<!--TIPO DE GRAFICA -->
                        data:{
                            labels:Fecha, //<!--DATO FECHA -->		
                            datasets:[{
                                    label:'TEMPERATURA CORPORAL',
                                    data:Temperatura, //<!--DATO TEMPERATURA -->
                                    backgroundColor: gradient,
                                    borderColor: "#000000",
                                    borderWidth:4,
                                        }]},
                        options:{
                            responsive: false,
                            title: {
                             display: true,
                             text: 'HISTORIAL DE LA TEMPERATURA CORPORAL'
                              },
                            
                            scales:{
                                yAxes:[{
                                    scaleLabel: {
                                        display: true,
                                        labelString: "TEMPERATURA [°C]",},
                                    ticks:{
                                        beginAtZero:true							
                                        }}],				
                                xAxes:[{
                                    scaleLabel: {
                                        display: true,
                                        labelString: "FECHA [dd/mm/aa]",
                                    }}]												
                        }}});	
                    document.getElementById('oli').style.display = 'block';
                    $("#DescTemp").click(function() {
                       myChart.options.title.text = 'Temperatura Corporal - Historial';
                       myChart.update({
                          duration: 0
                       });
                       var link = document.createElement('a');
                       link.href = myChart.toBase64Image();
                       link.download = 'Temperatura Corporal.png';
                       link.click();
                       myChart.options.title.text = 'HISTORIAL: TEMPERATURA CORPORAL';
                       myChart.update({
                          duration: 0
                       });
                    });	
                });			
                </script>
                
                <!-- GRAFICA PESO -->
                <!-- --------------------------------------------------------- -->
                <script>
                    var Pes = ""; var Fec = "";
                    var Peso =[]; var Fecha =[];
                      $("#btnGraficar").click(function() {
                        $(".GF").parent("tr").find(".GF").each(function() {
                         Fec +=$(this).html()+",";
                        });
                        $(".GP").parent("tr").find(".GP").each(function() {
                         Pes +=$(this).html()+",";
                        });			
                        Pes=Pes.split(',')
                        PesLen = Pes.length-1;
                        Fec=Fec.split(',')
                        for ( var i = 0; i < PesLen; i++ ) {
                            Peso[i]=Pes[i];
                            Fecha[i]=Fec[i];
                            }
                        var ctx= document.getElementById("GraPeso").getContext("2d");
                        let gradient=ctx.createLinearGradient(0,0,0,400);
                        gradient.addColorStop(0,"rgba(0,144,0,0.7)");
                        gradient.addColorStop(1,"rgba(0,200,0,0.1)");
                        var myChart= new Chart(ctx,{
                            type:"line", //<!--TIPO DE GRAFICA -->
                            data:{
                                labels:Fecha, //<!--DATO FECHA -->		
                                datasets:[{
                                        label:'PESO CORPORAL',
                                        data:Peso, //<!--DATO Peso -->
                                        backgroundColor: gradient,
                                        borderColor: "#000000",
                                        pointBackgroundColor: "#124500",
                                        pointBorderColor: "#124500",	
                                        borderWidth:4,
                                            }]},
                            options:{
                                responsive: false,
                                title: {
                                 display: true,
                                 text: 'HISTORIAL: PESO CORPORAL'
                                  },
                                scales:{
                                    yAxes:[{
                                        scaleLabel: {
                                            display: true,
                                            labelString: "PESO [Kg]",},
                                        ticks:{
                                            beginAtZero:true							
                                            }}],				
                                    xAxes:[{
                                        scaleLabel: {
                                            display: true,
                                            labelString: "FECHA [dd/mm/aa]",
                                        }}]												
                            }}});	
                        $("#DescPes").click(function() {
                           myChart.options.title.text = 'Peso Corporal - Historial';
                           myChart.update({
                              duration: 0
                           });
                           var link = document.createElement('a');
                           link.href = myChart.toBase64Image();
                           link.download = 'Peso Corporal.png';
                           link.click();
                           myChart.options.title.text = 'HISTORIAL: PESO CORPORAL';
                           myChart.update({
                              duration: 0
                           });
                        });		
                      });
                </script>

                <!-- GRAFICA ESTATURA -->
                <!-- --------------------------------------------------------- -->
                <script>
                    var Est = ""; var Fec = "";
                    var Estatura =[]; var Fecha =[];
                  $("#btnGraficar").click(function() {
                    $(".GF").parent("tr").find(".GF").each(function() {
                     Fec +=$(this).html()+",";
                    });
                    $(".GE").parent("tr").find(".GE").each(function() {
                     Est +=$(this).html()+",";
                    });		
                    Est=Est.split(',')
                    EstLen = Est.length-1;
                    Fec=Fec.split(',')
                    for ( var i = 0; i < EstLen; i++ ) {
                        Estatura[i]=Est[i];
                        Fecha[i]=Fec[i];
                        }
                    var ctx= document.getElementById("GraEstatura").getContext("2d");
                    let gradient=ctx.createLinearGradient(0,0,0,400);
                    gradient.addColorStop(0,"rgba(0,0,144,0.7)");
                    gradient.addColorStop(1,"rgba(0,0,200,0.1)");
                    var myChart= new Chart(ctx,{
                        type:"line", //<!--TIPO DE GRAFICA -->
                        data:{
                            labels:Fecha, //<!--DATO FECHA -->		
                            datasets:[{
                                    label:'ESTATURA CORPORAL',
                                    data:Estatura, //<!--DATO Estatura -->
                                    backgroundColor: gradient,
                                    borderColor: "#000000",
                                    pointBackgroundColor: "#124500",
                                    pointBorderColor: "#124500",
                                    borderWidth:4,					
                                        }]},
                        options:{
                            responsive: false,
                            title: {
                             display: true,
                             text: 'HISTORIAL: ESTATURA CORPORAL'
                              },
                            scales:{
                                yAxes:[{
                                    scaleLabel: {
                                        display: true,
                                        labelString: "ESTATURA [m]",},
                                    ticks:{
                                        beginAtZero:true							
                                        }}],				
                                xAxes:[{
                                    scaleLabel: {
                                        display: true,
                                        labelString: "FECHA [dd/mm/aa]",
                                    }}]												
                        }}});				
                    $("#DescEst").click(function() {
                       myChart.options.title.text = 'Estatura Corporal - Historial';
                       myChart.update({
                          duration: 0
                       });
                       var link = document.createElement('a');
                       link.href = myChart.toBase64Image();
                       link.download = 'Estatura Corporal.png';
                       link.click();
                       myChart.options.title.text = 'HISTORIAL: ESTATURA CORPORAL';
                       myChart.update({
                          duration: 0
                       });
                    });
                  });
                </script>
                

                <!-- GRAFICA IMC -->
                <!-- --------------------------------------------------------- -->
                <script>
                    var Im = ""; var Fec = "";
                    var IMC =[]; var Fecha =[];
                      $("#btnGraficar").click(function() {
                        $(".GF").parent("tr").find(".GF").each(function() {
                         Fec +=$(this).html()+",";
                        });
                        $(".GI").parent("tr").find(".GI").each(function() {
                         Im +=$(this).html()+",";
                        });			
                        Im=Im.split(',')
                        ImcLen = Im.length-1;
                        Fec=Fec.split(',')
                        for ( var i = 0; i < ImcLen; i++ ) {
                            IMC[i]=Im[i];
                            Fecha[i]=Fec[i];
                            }
                        var ctx= document.getElementById("GraImc").getContext("2d");
                        let gradient=ctx.createLinearGradient(0,0,0,400);
                        gradient.addColorStop(0,"rgba(200,0,255,0.8)");
                        gradient.addColorStop(1,"rgba(126, 2, 252, 0.3)");
                        var myChart= new Chart(ctx,{
                            type:"line", //<!--TIPO DE GRAFICA -->
                            data:{
                                labels:Fecha, //<!--DATO FECHA -->		
                                datasets:[{
                                        label:'INDICE DE MASA CORPORAL',
                                        data:IMC, //<!--DATO IMC -->
                                        backgroundColor:gradient,
                                        borderColor: "#000000",
                                        pointBackgroundColor: "#124500",
                                        pointBorderColor: "#124500",
                                        borderWidth:4,
                                            }]},
                            options:{
                                responsive: false,
                                title: {
                                 display: true,
                                 text: 'HISTORIAL: INDICE DE MASA CORPORAL'
                                  },
                                scales:{
                                    yAxes:[{
                                        scaleLabel: {
                                            display: true,
                                            labelString: "IMC [Kg/m^2]",},
                                        ticks:{
                                            beginAtZero:true							
                                            }}],				
                                    xAxes:[{
                                        scaleLabel: {
                                            display: true,
                                            labelString: "FECHA [dd/mm/aa]",
                                        }}]												
                            }}});	
                        $("#DescImc").click(function() {
                           myChart.options.title.text = 'IMC - Historial';
                           myChart.update({
                              duration: 0
                           });
                           var link = document.createElement('a');
                           link.href = myChart.toBase64Image();
                           link.download = 'IMC.png';
                           link.click();
                           myChart.options.title.text = 'HISTORIAL: INDICE DE MASA CORPORAL';
                           myChart.update({
                              duration: 0
                           });
                        });	
                      });
                </script>
                <!-- GRAFICA O2Sat -->
                <!-- --------------------------------------------------------- -->
                <script>
                    var O2S = "";  var Fec = "";
                    var O2Sat =[]; var Fecha =[];
                      $("#btnGraficar").click(function() {
                        $(".GF").parent("tr").find(".GF").each(function() {
                         Fec +=$(this).html()+",";
                        });
                        $(".GO").parent("tr").find(".GO").each(function() {
                         O2S +=$(this).html()+",";
                        });			
                        O2S=O2S.split(',')
                        O2SLen = O2S.length-1;
                        Fec=Fec.split(',')
                        for ( var i = 0; i < O2SLen; i++ ) {
                            O2Sat[i]=O2S[i];
                            Fecha[i]=Fec[i];
                            }
                        var ctx= document.getElementById("GraO2Sat").getContext("2d");
                        let gradient=ctx.createLinearGradient(0,0,0,400);
                        gradient.addColorStop(0,"rgba(255, 248, 8, 0.9)");
                        gradient.addColorStop(1,"rgba(255, 248, 8, 0.2)");
                        var myChart= new Chart(ctx,{
                            type:"line", //<!--TIPO DE GRAFICA -->
                            data:{
                                labels:Fecha, //<!--DATO FECHA -->		
                                datasets:[{
                                        label:'O2SAT',
                                        data:O2Sat, //<!--DATO O2Sat -->
                                        backgroundColor: gradient,
                                        borderColor: "#000000",
                                        pointBackgroundColor: "#124500",
                                        pointBorderColor: "#124500",	
                                            }]},
                            options:{
                                responsive: false,
                                title: {
                                 display: true,
                                 text: 'HISTORIAL: SATURACION DE OXIGENO EN LA SANGRE'
                                  },
                                scales:{
                                    yAxes:[{
                                        scaleLabel: {
                                            display: true,
                                            labelString: "O2SAT [%]",},
                                        ticks:{
                                            beginAtZero:true							
                                            }}],				
                                    xAxes:[{
                                        scaleLabel: {
                                            display: true,
                                            labelString: "FECHA [dd/mm/aa]",
                                        }}]												
                            }}});	
                        $("#DescO2Sat").click(function() {
                           myChart.options.title.text = 'O2Sat - Historial';
                           myChart.update({
                              duration: 0
                           });
                           var link = document.createElement('a');
                           link.href = myChart.toBase64Image();
                           link.download = 'O2SAT.png';
                           link.click();
                           myChart.options.title.text = 'HISTORIAL: SATURACION DE OXIGENO EN LA SANGRE';
                           myChart.update({
                              duration: 0
                           });
                        });				
                      });
                </script>	
                <!-- EXPORTAR TODOS LOS GRAFICOS -->
                <!-- ------------------------------------------------- -->
                <script>
                    $('#DescPDF').click(function(event) { 
                        var reportPageHeight = 1500;
                        var reportPageWidth = 2200;
                        var pdfCanvas = $('<canvas />').attr({
                            id: "canvaspdf",
                            width: reportPageWidth,
                            height: reportPageHeight
                        });	  
                          var pdfctx = $(pdfCanvas)[0].getContext('2d');
                          var pdfctxX = 0;
                          var pdfctxY = 0;
                          var buffer = 100;					  
                      $("canvas").each(function(index) {
                        var canvasHeight = $(this).innerHeight();
                        var canvasWidth = $(this).innerWidth();			
                        pdfctx.drawImage($(this)[0], pdfctxX, pdfctxY, canvasWidth, canvasHeight);
                        pdfctxX += canvasWidth + buffer;		
                        if (index % 2 === 1) {
                          pdfctxX = 0;
                          pdfctxY += canvasHeight + buffer;
                            }
                         });
                        var pdf = new jsPDF('l', 'pt', [reportPageWidth, reportPageHeight]);
                        pdf.addImage($(pdfCanvas)[0], 'PNG', 0, 0);
                        pdf.save('Reporte Grafico IoMT.pdf');
                    });
                    function done(){
                      alert("haha");
                      var url=myLine.toBase64Image();
                      document.getElementById("url").src=url;
                    }
                </script> 	
            </html>        
            """)
            archivo1.close()

            wb = openpyxl.Workbook()
            ws=wb.active
            temp=(["Nombre", personName])
            temp2=(["Cédula", cedula])
            temp3=(["Curso", curso])
            temp4=(["Fecha de Nacimiento", nacimiento])
            temp5=(["Edad", edad])
            temp7=(["Fecha","Hora", "Temperatura","Peso","Altura","IMC","O2Sat"])
            ws.append(temp)
            ws.append(temp2)
            ws.append(temp3)
            ws.append(temp4)
            ws.append(temp5)
            ws.append(temp7)
            wb.save('/home/pi/IoMTv2/DatosExcel/'+personName+".xlsx")
            showMessage('Registro Realizado Correctamente', type='info', timeout=2000)
            
            entradanombre.delete(0,END) 
            entradaapellido.delete(0,END) 
            entradacedula.delete(0,END) 
            entradacurso.delete(0,END)
            entradaaño.delete(0,END)
            entradames.delete(0,END)
            entradadia.delete(0,END) 
        
    botonaceptar = tkinter.Button(ventana2, text = 'Aceptar', command  = capturarrostro, padx=74, pady=18,font=('ALGERIAN 20 bold') )
    botonaceptar.place(x=291, y=601)
    botonregresar = tkinter.Button(ventana2, text = 'Regresar', command  = atras , padx=66, pady=18, font=('ALGERIAN 20 bold'))
    botonregresar.place(x=722, y=601)
    

def actualizargithub():    
    repo = Repo('/home/pi/IoMTv2/daniels36999.github.io')  # if repo is CWD just do '.'
    repo.index.add(['/home/pi/IoMTv2/daniels36999.github.io/Datos Almacenados'])
    repo.index.commit('actualizando')
    origin = repo.remote('origin')
    origin.push()
    
def reconocimientof():
    tvX1=40; tvX2=190
    etX1=130; etY1=30
    btX1=200; btY1=50

    Temperaturaz=0.0
    Pesoz=0.0
    Estaturaz=0.0
    Imcz=0.0
    O2Satz=0.0

    ventIoMT = Toplevel(ventana)
    ventIoMT.title("IoMT")
    ventIoMT.geometry('%dx%d+%d+%d' % (360, 450, 51, 60))
    ventIoMT.resizable(0, 0)

    # VIZUALIZACION DE RESULTADO
    tv1Temperatura= tkinter.Label(ventIoMT, text= str("Temperatura"), font=("Times New Roman",20))
    tv1Temperatura.place(x=tvX1, y=100)
    tv1Peso= tkinter.Label(ventIoMT, text= str("Peso"), font=("Times New Roman",20))
    tv1Peso.place(x=tvX1, y=160)
    tv1Estatura= tkinter.Label(ventIoMT, text= str("Estatura"), font=("Times New Roman",20))
    tv1Estatura.place(x=tvX1, y=220)
    tv1Imc= tkinter.Label(ventIoMT, text= str("IMC"), font=("Times New Roman",20))
    tv1Imc.place(x=tvX1, y=280)
    tv1O2Sat= tkinter.Label(ventIoMT, text= str("O2Sat"), font=("Times New Roman",20))
    tv1O2Sat.place(x=tvX1, y=340)

    # VIZUALIZACION DE RESULTADO
    tv2Temperatura= tkinter.Label(ventIoMT, text= str(Temperaturaz), font=("Times New Roman",20),bg='#fff')
    tv2Temperatura.place(x=tvX2, y=100, width=etX1, height=etY1)
    tv2Peso= tkinter.Label(ventIoMT, text= str(Pesoz), font=("Times New Roman",20),bg='#fff')
    tv2Peso.place(x=tvX2, y=160, width=etX1, height=etY1)
    tv2Estatura= tkinter.Label(ventIoMT, text= str(Estaturaz), font=("Times New Roman",20),bg='#fff')
    tv2Estatura.place(x=tvX2, y=220, width=etX1, height=etY1)
    tv2Imc= tkinter.Label(ventIoMT, text= str(Imcz), font=("Times New Roman",20),bg='#fff')
    tv2Imc.place(x=tvX2, y=280, width=etX1, height=etY1)
    tv2O2Sat= tkinter.Label(ventIoMT, text= str(O2Satz), font=("Times New Roman",20),bg='#fff')
    tv2O2Sat.place(x=tvX2, y=340, width=etX1, height=etY1)
    
    serialArduino = serial.Serial('/dev/ttyACM0',9600)
    
    #RECONOCIMIENTO FACIAL
    dataPath = '/home/pi/IoMTv2/Data' #Cambia a la ruta donde hayas almacenado Data
    imagePaths = os.listdir(dataPath)
    print('imagePaths=',imagePaths)

    face_recognizer = cv2.face.EigenFaceRecognizer_create()
    face_recognizer.read('modeloEigenFacepruebainterface.xml')
    cap = cv2.VideoCapture(0,cv2.CAP_V4L)
    faceClassif = cv2.CascadeClassifier(cv2.data.haarcascades+'haarcascade_frontalface_default.xml')
    
    count1 = 0
    count2 = 0
    
    while True:
        ret,frame = cap.read()
        if ret == False: break
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        auxFrame = gray.copy()
        faces = faceClassif.detectMultiScale(gray,1.3,5)       
        for (x,y,w,h) in faces:
            rostro = auxFrame[y:y+h,x:x+w]
            rostro = cv2.resize(rostro,(150,150),interpolation= cv2.INTER_CUBIC)
            result = face_recognizer.predict(rostro)
            cv2.putText(frame,'{}'.format(result),(x,y-5),1,1.3,(255,255,0),1,cv2.LINE_AA)
            if result[1] < 5000:
                cv2.putText(frame,'{}'.format(imagePaths[result[0]]),(x,y-25),2,1.1,(0,255,0),1,cv2.LINE_AA)
                cv2.rectangle(frame, (x,y),(x+w,y+h),(0,255,0),2)
                nombrefinal = '{}'.format(imagePaths[result[0]])
                nombreparallenar = str(nombrefinal)
                count1 = count1 + 1
            else:
                cv2.putText(frame,'Desconocido',(x,y-20),2,0.8,(0,0,255),1,cv2.LINE_AA)
                cv2.rectangle(frame, (x,y),(x+w,y+h),(0,0,255),2)
                
        cv2.imshow('frame',frame)    
        print(count1)
        count2=count2+1
        
        k =  cv2.waitKey(1)
        if k == 27 or count1 >= 10:     
            break
        if count2 >= 300:         
            break
            
    cap.release()
    cv2.destroyAllWindows()
    datos=pd.read_csv('/home/pi/IoMTv2/Todos los Datos/'+nombreparallenar+'.csv', header= None)
    df=pd.DataFrame(datos)
    a = str(df.iat[2,1])
    ced = str(df.iat[1,1])
    fec = str(df.iat[3,1])
    edad1 = str(df.iat[4,1])
    
    cont1=0
    cont2=0
    contadorprogres1=0
    showMessage('Se Recolectarán los Datos al cerrarse este mensaje', type='info', timeout=2000)
    
    while True :
        cad=serialArduino.readline().decode('ascii')
        comando='O'
        comandoBytes = comando.encode()
        serialArduino .write(comandoBytes)
        progress1['value'] = contadorprogres1
        txt1['text']=round(progress1['value']),'%'
        ventana.update_idletasks()      
   
        try:
            if(cont1>=10):
                cad=serialArduino.readline().decode('ascii')
                datoss=cad.splitlines()
                d0=str(datoss[0])
                d1=d0.replace("b","")
                d2=d1.replace("'","")
                d3=d2.split(",")
                dato1=str(d3[0])
                dato2=str(d3[1])
                dato3=str(d3[2])
                dato4=str(d3[3])
                dato5=str(d3[4])               
                dato9=float(dato3)
                if(dato9<1):
                    dato9=1
                    
                dato6=float(dato2)/(float(dato9)*float(dato9))
                dato7=round(dato6,2)
                dato8=str(dato7)
                
                tv2Temperatura['text']=dato1
                tv2Peso['text']=dato2
                tv2Estatura['text']=dato3
                tv2Imc['text']=dato8
                tv2O2Sat['text']=dato5
                
                print(dato1,dato2,dato3,dato8,dato5)
                print("--------------------")
                cont1=cont1+1
                ahora=datetime.datetime.now()
                ahora1=ahora.strftime("%d/%m/%y")
                ahora2=ahora.strftime("%Hh/%Mm/%Ss")
                ahora3=ahora.strftime("%d-%m-%y")
                ahora4=ahora.strftime("%H-%M-%S")
                client = mqtt.Client()
                client.connect("test.mosquitto.org",1883,60)
                client.publish("Topic1/Topic2/Topic3", str(dato1)+';'+str(dato2)+';'+str(dato3)+';'+str(dato5)+';'+str(dato5));
                client.disconnect();
     
            if(cont2==20):
                
                comando='F'
                comandoBytes = comando.encode()
                serialArduino.write(comandoBytes)
                with open('/home/pi/IoMTv2/daniels36999.github.io/Datos Almacenados/'+a+'/'+nombreparallenar+'.html',"r") as f:
                    newline=[]
                    for word in f.readlines():
                        newline.append(word.replace("<!-- AGREGAR TABLA-->","""<tr><td class="GF">"""+ahora1+"""</td><td class="GH">"""+ahora2+"""</td><td class="GT">"""+dato1+"""</td><td class="GP">"""+dato2+"""</td><td class="GE">"""+dato3+"""</td><td class="GI">"""+str(dato8)+"""</td><td class="GO">"""+dato5+"""</td></tr>
                        <!-- AGREGAR TABLA-->"""))  ## Replace the keyword while you copy.
                with open('/home/pi/IoMTv2/daniels36999.github.io/Datos Almacenados/'+a+'/'+nombreparallenar+'.html',"w") as f:
                    for line in newline:
                        f.writelines(line)
                    f.close()
 
                wb2=openpyxl.load_workbook('//home/pi/IoMTv2/DatosExcel/'+nombreparallenar+'.xlsx')
                ws=wb2.active
                temp=([ahora1, ahora2, dato1,dato2,dato3,str(dato8),dato5])
                ws.append(temp)
                wb2.save('/home/pi/IoMTv2/DatosExcel/'+nombreparallenar+'.xlsx')
                client = mqtt.Client()
                client.connect("test.mosquitto.org",1883,60)
                #INGRESE LOS TOPICS QUE TENGA HABILITADO PARA IoMT
                client.publish("Topic1/Topic2/Topic3", str(dato1)+';'+str(dato2)+';'+str(dato3)+';'+str(dato8)+';'+str(dato5));
                client.disconnect();
                try:
                    conexion = mariadb.connect(
                        user="Usuario", #INGRESE EL USUARIO DE SU BD
                        password="********",#INGRESE LA CONTRASEÑA DE SU BD
                        host="127.0.0.1",
                        port=3306,
                        database="prueba"
                       )
                   
                    print("CONECTADO")
                    cursor = conexion.cursor()
                   
                except mariadb.Error as error:
                    print(f"Error al conectar con la base de datos: {error}")
                    sys.exit(1)

                try:
                    cursor.execute("INSERT INTO iomt"
                    "(Fecha,Hora,Nombres,Cédula,Curso,FechaNacimiento,Edad,Temperatura,Peso,Altura,IMC,O2Sat)"
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?,?)",
                    (ahora3, ahora4, nombreparallenar, ced, a ,fec,edad1, str(dato1),str(dato2),str(dato3),str(dato8),str(dato5)))
                    conexion.commit()
                    cursor.close()
                    conexion.close
                except mariadb.Error as error_registro:
                    print(f"Error en el registro: {error_registro}")

                cont1=0
                cont2=0
                progress1['value'] = 0
                txt1['text']=round(progress1['value']),'%'
                actualizargithub()
                showMessage('Datos Almacenados Correctamente', type='info', timeout=2000)
                ventIoMT.destroy()
                break
            
            cont1=cont1+1
            cont2=cont2+1
            print(cont1,cont2)
            contadorprogres1=((cont2*100)/20)
            time.sleep(0.5)
        except IndexError:
            showMessage('Vuelva a Intentarlo', type='error', timeout=2000)
    client = mqtt.Client()
    client.connect("test.mosquitto.org",1883,60)
    client.publish("Topic1/Topic2/Topic3", str("00:00")+';'+str("00:00")+';'+str("00:00")+';'+str("00:00")+';'+str("00:00"));
    client.disconnect();
    ventana.deiconify() 

ventana = tkinter.Tk()
ventana.title('Ventana Principal')
ventana.attributes('-fullscreen', True)

im = PIL.Image.open("/home/pi/IoMTv2/fondos/inicio.png")
im1 = PIL.Image.open("/home/pi/IoMTv2/fondos/registro.png")
photo = PIL.ImageTk.PhotoImage(im)
photo1 = PIL.ImageTk.PhotoImage(im1)
label = Label(ventana, image=photo).place(x=0,y=0,relwidth=1.0,relheight=1.0)

botonregistro = tkinter.Button(ventana, text = 'Registro', command  = ventanaregis,bg='#37EF31', font=('ALGERIAN 20 bold'))
botonregistro.place(x=100 , y=100, width=250,height=110)


botonautentificacion = tkinter.Button(ventana, text = 'Autentificación', command = reconocimientof,bg='#20E51A',font=('ALGERIAN 18 bold'))
botonautentificacion.place(x=100 , y=250, width=250,height=110)
botoniomt = tkinter.Button(ventana, text = 'IOMT \n Usuarios no registrados', command = inicio,bg='#17B712',font=('ALGERIAN 12 bold'))
botoniomt.place(x=100, y=390, width=250,height=110)
progress1 = Progressbar(ventana, orient = HORIZONTAL,length = 100, mode = 'determinate')
progress1.place(x=440, y=270, width=700,height=60)
progress = Progressbar(ventana, orient = HORIZONTAL,length = 100, mode = 'determinate')
progress.place(x=440, y=420, width=700,height=60)
txt1 = Label(ventana,text = '0%',font=("Times New Roman",40))
txt1.place(x=1150, y=270 )
txt2 = Label(ventana,text = '0%',font=("Times New Roman",40))
txt2.place(x=1150, y=420 )
serialArduino = serial.Serial('/dev/ttyACM0',9600)
ventana.mainloop()e